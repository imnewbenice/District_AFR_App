import streamlit as st
from openpyxl import load_workbook, Workbook
from io import BytesIO
import os

st.title("District AFR Data Extractor and Template Filler")

# Define the cell mappings based on the provided template layout
cell_mapping = {
    "E11": "Regular Education", "G11": "Regular Education",
    "E12": "Special Education", "G12": "Special Education",
    "E13": "Pupil Transportation", "G13": "Pupil Transportation",
    "E14": "Desegregation", "G14": "Desegregation",
    "E15": "Dropout Prevention Programs", "G15": "Dropout Prevention Programs",
    "E16": "Joint Career & Tech. Ed. & Voc. Ed. Center", "G16": "Joint Career & Tech. Ed. & Voc. Ed. Center",
    "E17": "K-3 Reading Program", "G17": "K-3 Reading Program", "I17": "General",
    # Add all other cells according to the provided template up to B69 and I69
    "B18": "Maintenance and Operation Total", "C18": "Maintenance and Operation Total", "D18": "Maintenance and Operation Total",
    "E18": "Maintenance and Operation Total", "G18": "Maintenance and Operation Total", "I18": "Special Revenue",
    # ... (extend this for each row and cell in the provided template)
}

# Upload AFR files and the template file
afr_files = st.file_uploader("Upload AFR Workbooks (2017-2024)", type="xlsx", accept_multiple_files=True)
template_file = st.file_uploader("Upload AFR Summary Template", type="xlsx")

# Start processing once all files are uploaded and button is clicked
if st.button("Start Processing") and afr_files and template_file:
    # Initialize a progress bar
    progress_bar = st.progress(0)
    total_files = len(afr_files)
    
    # Load the template workbook to populate with data
    template_wb = load_workbook(template_file, data_only=True)
    
    # Initialize a new workbook to store filled templates
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove the default blank sheet

    # Process each uploaded AFR file
    for i, afr_file in enumerate(afr_files):
        # Extract the name of the file without extension for sheet naming
        file_name = os.path.splitext(afr_file.name)[0]

        # Load the AFR workbook
        afr_wb = load_workbook(afr_file, data_only=True)

        # Check if "Summary" sheet exists in the workbook
        if "Summary" in afr_wb.sheetnames:
            afr_summary_sheet = afr_wb["Summary"]
            
            # Create a copy of the template sheet to populate for this AFR file
            new_sheet = output_wb.create_sheet(title=file_name)
            template_sheet = template_wb.active  # Use the first sheet in the template as base

            # Copy the entire template sheet structure to the new sheet
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value

            # Populate the new sheet with extracted data
            for cell_ref, description in cell_mapping.items():
                # Extract value from the specified cell in the AFR's "Summary" sheet
                if cell_ref in afr_summary_sheet:
                    cell_value = afr_summary_sheet[cell_ref].value
                    if cell_value is not None:
                        new_sheet[cell_ref] = cell_value

        # Update progress bar after processing each file
        progress_bar.progress((i + 1) / total_files)

    # Save the combined workbook to a BytesIO buffer for download
    output = BytesIO()
    output_wb.save(output)
    output.seek(0)

    # Provide a download button for the generated combined template
    st.download_button(
        label="Download Filled AFR Summary Template",
        data=output,
        file_name="AFR_Summary_Combined_Template_Filled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Display completion message
    st.success("Data extraction and template filling complete!")
